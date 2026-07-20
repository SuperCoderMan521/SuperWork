#!/usr/bin/env python3
"""Generate a professional order Excel file with multiple sheets."""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import random
import os

wb = Workbook()

# ── Color Palette & Styles ──
HEADER_FILL = PatternFill(start_color="1A2A3A", end_color="1A2A3A", fill_type="solid")
HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
ACCENT_FILL = PatternFill(start_color="00D4AA", end_color="00D4AA", fill_type="solid")
ACCENT_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="1A2A3A")
TITLE_FONT  = Font(name="Microsoft YaHei", size=16, bold=True, color="1A2A3A")
SUBTITLE_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="1A2A3A")
DATA_FONT   = Font(name="Microsoft YaHei", size=10)
NUM_FONT    = Font(name="Microsoft YaHei", size=10, color="D77757")
TOTAL_FONT  = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
TOTAL_FILL  = PatternFill(start_color="D77757", end_color="D77757", fill_type="solid")
LIGHT_FILL  = PatternFill(start_color="F5F8FA", end_color="F5F8FA", fill_type="solid")
WHITE_FILL  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCDDEE"),
    right=Side(style="thin", color="CCDDEE"),
    top=Side(style="thin", color="CCDDEE"),
    bottom=Side(style="thin", color="CCDDEE"),
)
BOTTOM_BORDER = Border(
    bottom=Side(style="medium", color="1A2A3A"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right", vertical="center")
NUM_FMT = '#,##0.00'
QTY_FMT = '#,##0'

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

def style_data_row(ws, row, max_col, alt=False):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = DATA_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        if alt:
            cell.fill = LIGHT_FILL
        else:
            cell.fill = WHITE_FILL

# =====================================================
# Sheet 1 — 订单列表
# =====================================================
ws1 = wb.active
ws1.title = "订单列表"
ws1.sheet_properties.tabColor = "00D4AA"

# Column widths
col_widths = [6, 18, 14, 12, 12, 10, 14, 14, 10, 14, 20]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Title
ws1.merge_cells("A1:K1")
title_cell = ws1.cell(row=1, column=1, value="📦  订单管理列表")
title_cell.font = TITLE_FONT
title_cell.alignment = Alignment(horizontal="left", vertical="center")
ws1.row_dimensions[1].height = 40

ws1.merge_cells("A2:K2")
ws1.cell(row=2, column=1, value=f"生成日期：{datetime.now().strftime('%Y-%m-%d %H:%M')}   |   共 20 条记录"
        ).font = Font(name="Microsoft YaHei", size=9, color="8899AA")

# Headers (row 4)
headers = ["序号", "订单号", "客户名称", "联系电话", "订单日期", "状态",
           "商品名称", "数量", "单价(元)", "金额(元)", "备注"]
for col, h in enumerate(headers, 1):
    ws1.cell(row=4, column=col, value=h)
style_header_row(ws1, 4, len(headers))

# Sample data
products = [
    ("iPhone 16 Pro Max", 8999),
    ("MacBook Air M4", 7999),
    ("AirPods Pro 3", 1999),
    ("iPad Air 13\"", 5499),
    ("Apple Watch S10", 3499),
    ("Magic Keyboard", 999),
    ("USB-C 充电器 35W", 399),
    ("AirTag 四件装", 799),
    ("Mac mini M4 Pro", 12499),
    ("Apple Vision Pro", 29999),
]
statuses = ["已支付", "已发货", "已完成", "待支付", "已取消"]
names = ["张伟", "李娜", "王强", "赵敏", "陈小明", "刘洋", "杨静", "黄磊",
         "周杰", "吴芳", "徐峰", "孙莉", "马超", "朱婷", "胡亮", "林萍"]

for i in range(1, 21):
    row = i + 4
    pid = f"ORD-{datetime.now().year}{datetime.now().month:02d}{i:04d}"
    customer = random.choice(names)
    phone = f"138{random.randint(10000000, 99999999)}"
    date = (datetime.now() - timedelta(days=random.randint(0, 60))).strftime("%Y-%m-%d")
    status = random.choice(statuses)
    prod_name, unit_price = random.choice(products)
    qty = random.randint(1, 10)
    amount = unit_price * qty
    remark = random.choice(["", "加急", "企业采购", "礼品包装", "分期付款"])

    row_data = [i, pid, customer, phone, date, status,
                prod_name, qty, unit_price, amount, remark]
    for col, val in enumerate(row_data, 1):
        cell = ws1.cell(row=row, column=col, value=val)
    style_data_row(ws1, row, len(headers), alt=(i % 2 == 0))

    # Special formatting
    ws1.cell(row=row, column=8).number_format = QTY_FMT
    ws1.cell(row=row, column=9).number_format = NUM_FMT
    ws1.cell(row=row, column=10).number_format = NUM_FMT
    ws1.cell(row=row, column=10).font = NUM_FONT

    # Status color
    status_cell = ws1.cell(row=row, column=6)
    if status == "已支付":
        status_cell.font = Font(name="Microsoft YaHei", size=10, color="4CAF50")
    elif status == "已发货":
        status_cell.font = Font(name="Microsoft YaHei", size=10, color="2196F3")
    elif status == "已完成":
        status_cell.font = Font(name="Microsoft YaHei", size=10, color="9E9E9E")
    elif status == "待支付":
        status_cell.font = Font(name="Microsoft YaHei", size=10, color="FF9800")
    elif status == "已取消":
        status_cell.font = Font(name="Microsoft YaHei", size=10, color="F44336")

# Summary row
summary_row = 4 + 20 + 1
ws1.merge_cells(f"A{summary_row}:G{summary_row}")
total_label = ws1.cell(row=summary_row, column=1, value="合  计")
total_label.font = TOTAL_FONT
total_label.fill = TOTAL_FILL
total_label.alignment = CENTER
total_label.border = THIN_BORDER
for c in range(2, 8):
    cell = ws1.cell(row=summary_row, column=c)
    cell.fill = TOTAL_FILL
    cell.border = THIN_BORDER

total_qty = sum(ws1.cell(row=r, column=8).value for r in range(5, 25))
total_amt = sum(ws1.cell(row=r, column=10).value for r in range(5, 25))

for col, val in [(8, total_qty), (9, ""), (10, total_amt)]:
    cell = ws1.cell(row=summary_row, column=col, value=val)
    cell.fill = TOTAL_FILL
    cell.font = TOTAL_FONT
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    if col == 8:
        cell.number_format = QTY_FMT
    elif col == 10:
        cell.number_format = NUM_FMT

for c in range(11, len(headers) + 1):
    cell = ws1.cell(row=summary_row, column=c)
    cell.fill = TOTAL_FILL
    cell.border = THIN_BORDER

# Freeze panes
ws1.freeze_panes = "A5"

# Auto filter
ws1.auto_filter.ref = f"A4:K{4 + 20}"

# =====================================================
# Sheet 2 — 订单详情（单笔订单模板）
# =====================================================
ws2 = wb.create_sheet("订单详情")
ws2.sheet_properties.tabColor = "D77757"

col_widths2 = [4, 20, 20, 14, 14, 14, 20]
for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# Title
ws2.merge_cells("A1:G1")
ws2.cell(row=1, column=1, value="订  单  详  情").font = TITLE_FONT
ws2.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 40

# Order info block
info = [
    ("订单号", "ORD-20250719-0001"),
    ("客户名称", "张伟"),
    ("联系电话", "13888886666"),
    ("订单日期", "2025-07-18"),
    ("订单状态", "已支付"),
    ("收货地址", "北京市朝阳区建国路88号SOHO现代城A座1508"),
]

for i, (label, val) in enumerate(info):
    row = 3 + i
    ws2.merge_cells(f"A{row}:B{row}")
    label_cell = ws2.cell(row=row, column=1, value=label)
    label_cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color="8899AA")
    label_cell.alignment = LEFT

    ws2.merge_cells(f"C{row}:G{row}")
    val_cell = ws2.cell(row=row, column=3, value=val)
    val_cell.font = Font(name="Microsoft YaHei", size=10, color="1A2A3A")
    val_cell.alignment = LEFT

    if i == 4:  # status
        val_cell.font = Font(name="Microsoft YaHei", size=10, color="4CAF50", bold=True)

# Detail table header
detail_headers = ["序号", "商品编号", "商品名称", "单价(元)", "数量", "小计(元)", "备注"]
detail_start = 11
for col, h in enumerate(detail_headers, 1):
    ws2.cell(row=detail_start, column=col, value=h)
style_header_row(ws2, detail_start, len(detail_headers))

# Detail items
detail_items = [
    ("P001", "iPhone 16 Pro Max 256GB", 8999, 2),
    ("P002", "AirPods Pro 3 USB-C", 1999, 1),
    ("P003", "AppleCare+ iPhone 两年", 1499, 2),
    ("P004", "MagSafe 充电器", 399, 1),
    ("P005", "20W USB-C 充电器", 149, 3),
]

for i, (code, name, price, qty) in enumerate(detail_items):
    row = detail_start + 1 + i
    subtotal = price * qty
    row_data = [i + 1, code, name, price, qty, subtotal, ""]
    for col, val in enumerate(row_data, 1):
        ws2.cell(row=row, column=col, value=val)
    style_data_row(ws2, row, len(detail_headers), alt=(i % 2 == 0))
    ws2.cell(row=row, column=4).number_format = NUM_FMT
    ws2.cell(row=row, column=5).number_format = QTY_FMT
    ws2.cell(row=row, column=6).number_format = NUM_FMT
    ws2.cell(row=row, column=6).font = NUM_FONT

# Summary
sum_start = detail_start + 1 + len(detail_items)
merge_labels = ["", "", "", "", "", ""]

summary_data = [
    ("商品总数", "", "", "", "", 9),
    ("商品总金额", "", "", "", "", 8999 * 2 + 1999 + 1499 * 2 + 399 + 149 * 3),
]

for i, (label, *_) in enumerate(summary_data):
    row = sum_start + i
    ws2.merge_cells(f"A{row}:E{row}")
    label_cell = ws2.cell(row=row, column=1, value=label)
    label_cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color="8899AA")
    label_cell.alignment = RIGHT
    label_cell.border = THIN_BORDER
    for c in range(2, 6):
        ws2.cell(row=row, column=c).border = THIN_BORDER

subtotal_val = 8999 * 2 + 1999 + 1499 * 2 + 399 + 149 * 3
# Actually let me compute properly
subtotal = sum(price * qty for _, _, price, qty in detail_items)
discount = 500
tax = round(subtotal * 0.13, 2)
grand_total = subtotal - discount + tax

finance_rows = [
    ("小计", subtotal),
    ("优惠折扣", -discount),
    ("税率 (13%)", tax),
    ("总计", grand_total),
]

for i, (label, val) in enumerate(finance_rows):
    row = sum_start + 2 + i
    ws2.merge_cells(f"A{row}:E{row}")
    label_cell = ws2.cell(row=row, column=1, value=label)
    val_cell = ws2.cell(row=row, column=6, value=val)
    val_cell.number_format = NUM_FMT

    if i == 3:  # 总计行
        label_cell.font = TOTAL_FONT
        label_cell.fill = TOTAL_FILL
        val_cell.font = TOTAL_FONT
        val_cell.fill = TOTAL_FILL
        for c in range(2, 6):
            ws2.cell(row=row, column=c).fill = TOTAL_FILL
        for c in range(1, len(detail_headers) + 1):
            ws2.cell(row=row, column=c).border = THIN_BORDER
        ws2.cell(row=row, column=6).fill = TOTAL_FILL
    else:
        label_cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color="1A2A3A")
        val_cell.font = Font(name="Microsoft YaHei", size=10, bold=True)
        for c in range(1, len(detail_headers) + 1):
            ws2.cell(row=row, column=c).border = THIN_BORDER

    label_cell.alignment = RIGHT
    val_cell.alignment = CENTER

# =====================================================
# Sheet 3 — 数据看板
# =====================================================
ws3 = wb.create_sheet("数据看板")
ws3.sheet_properties.tabColor = "5769F7"

# This sheet provides a summary dashboard
col_widths3 = [22, 16, 16, 16, 16, 16]
for i, w in enumerate(col_widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.merge_cells("A1:F1")
ws3.cell(row=1, column=1, value="📊  订单数据看板").font = TITLE_FONT
ws3.row_dimensions[1].height = 40

# KPI cards
kpi_data = [
    ("总订单数", "128 单"),
    ("本月订单", "36 单"),
    ("总营业额", "¥ 685,432"),
    ("本月营收", "¥ 187,650"),
    ("客单价", "¥ 5,355"),
    ("退货率", "2.3%"),
]

for i, (label, val) in enumerate(kpi_data):
    col_start = (i % 3) * 2 + 1
    row = 3 + (i // 3) * 3

    ws3.merge_cells(start_row=row, start_column=col_start,
                     end_row=row, end_column=col_start + 1)
    cell = ws3.cell(row=row, column=col_start, value=label)
    cell.font = Font(name="Microsoft YaHei", size=10, color="8899AA")
    cell.alignment = Alignment(horizontal="center", vertical="bottom")
    cell.fill = LIGHT_FILL
    for c in range(col_start, col_start + 2):
        ws3.cell(row=row, column=c).fill = LIGHT_FILL
        ws3.cell(row=row, column=c).border = BOTTOM_BORDER

    ws3.merge_cells(start_row=row + 1, start_column=col_start,
                     end_row=row + 1, end_column=col_start + 1)
    cell = ws3.cell(row=row + 1, column=col_start, value=val)
    cell.font = Font(name="Microsoft YaHei", size=22, bold=True, color="1A2A3A")
    cell.alignment = Alignment(horizontal="center", vertical="top")
    cell.fill = LIGHT_FILL
    for c in range(col_start, col_start + 2):
        ws3.cell(row=row + 1, column=c).fill = LIGHT_FILL
        ws3.cell(row=row + 1, column=c).border = BOTTOM_BORDER

# Monthly trend table
ws3.merge_cells("A10:F10")
ws3.cell(row=10, column=1, value="月度销售趋势（2025年）").font = SUBTITLE_FONT
ws3.row_dimensions[10].height = 30

month_headers = ["月份", "订单数", "营业额(元)", "成本(元)", "利润(元)", "利润率"]
for col, h in enumerate(month_headers, 1):
    ws3.cell(row=11, column=col, value=h)
style_header_row(ws3, 11, len(month_headers))

monthly_data = [
    ("1月", 28, 142000, 106500, 35500, "25.0%"),
    ("2月", 22, 108000, 81000, 27000, "25.0%"),
    ("3月", 35, 178000, 133500, 44500, "25.0%"),
    ("4月", 30, 152000, 114000, 38000, "25.0%"),
    ("5月", 32, 165000, 123750, 41250, "25.0%"),
    ("6月", 38, 195000, 146250, 48750, "25.0%"),
]

for i, (month, orders, revenue, cost, profit, margin) in enumerate(monthly_data):
    row = 12 + i
    row_data = [month, orders, revenue, cost, profit, margin]
    for col, val in enumerate(row_data, 1):
        ws3.cell(row=row, column=col, value=val)
    style_data_row(ws3, row, len(month_headers), alt=(i % 2 == 0))
    ws3.cell(row=row, column=3).number_format = NUM_FMT
    ws3.cell(row=row, column=4).number_format = NUM_FMT
    ws3.cell(row=row, column=5).number_format = NUM_FMT

# ── Save ──
desktop = os.path.expanduser("~/Desktop")
output_path = os.path.join(desktop, "订单管理系统.xlsx")
wb.save(output_path)
print(f"[OK] Excel saved to: {output_path}")